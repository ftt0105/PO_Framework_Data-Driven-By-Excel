from openpyxl import *
import os
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill
from openpyxl.styles import  colors,Font, Border,Side

class ExcelUtil():
    def __init__(self,excel_file_path):
        if os.path.exists(excel_file_path):
            self.excel_file_path = excel_file_path
            #加载excel文件
            self.wb = load_workbook(self.excel_file_path)
            #获取当前文件默认 sheet 的对象
            self.sheet = self.wb.active  #
        else:
            print("Excel 文件 %s 路径不存在" %excel_file_path)
            self.excel_file_path = None
            self.wb = None
            self.sheet = None

    #获取excel所有的sheet名称
    def get_sheet_names(self):
        if self.wb is not None:
            return self.wb.sheetnames
        else:
            return None

    #获取当前操作excel文件的绝对路径
    def get_file_path(self):
        return self.excel_file_path

    #设置当前操作的excel sheet对象是哪个，用序号来进行设置
    def set_sheet_by_index(self,index):#序号从0开始
        if not isinstance(index,int):
            print("您设定的sheet 序号不是数字类型，请重新设定")
            return None
        if not 0<=index<len(self.get_sheet_names()):
            print("您设定的sheet 序号不存在，请重新设定！最大sheet编号是%s" %(len(self.get_sheet_names())-1))
            return None
        self.sheet = self.wb[self.get_sheet_names()[index]]
        return self.sheet

    # 设置当前操作的excel sheet对象是哪个，用名字来进行设置
    def set_sheet_by_name(self,sheet_name):
        if not sheet_name in self.get_sheet_names():
            print("您设定的sheet名称%s不存在，请重新选择sheet名字" %sheet_name)
            return
        self.sheet = self.wb[sheet_name]
        return self.sheet

    #创建新的sheet
    def create_sheet(self,sheet_name):
        if sheet_name in self.get_sheet_names():
            print("sheet 名称%s已经存在，请重新选择个sheet名字" %sheet_name)
            return
        self.wb.create_sheet(sheet_name)
        self.wb.save(self.get_file_path())

    #获取当前excel文件中最大的行数
    def get_max_row_count(self):
        return self.sheet.max_row

    # 获取当前excel文件中最大的列数
    def get_max_col_count(self):
        return self.sheet.max_column

    #获得当前sheet中所有的单元格对象，返回一个带有多个子列表的列表
    def get_sheet_all_cells(self):
        cell_objects = []#存储整个儿sheet的单元对象列表，每一行cell会存在一个子列表中
        for row in self.sheet.iter_rows():
            row_cell_objects = []#存储某一行的所有cell对象
            for cell in row:
                row_cell_objects.append(cell)
            cell_objects.append(row_cell_objects)
        return cell_objects

    #获取当前sheet中所有的单元格对象的值，返回一个带有多个子列表的列表
    def get_sheet_all_cell_values(self):
        values = []
        for row in self.get_sheet_all_cells():
            row_values = []
            for cell in row:
                row_values.append(cell.value)
            values.append(row_values)
        return values

    #获取一个单元格的值，行号从0开始，列号从0开始
    def get_cell_value(self,row_no,col_no):
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            print("输入的行号%s,或者输入的列表%s不是整数" %(row_no,col_no))
            return  None
        if not 0<=row_no<self.get_max_row_count():
            print("输入的行号不在合法的行号范围内，需要大于等于0且小于最大行号的范围内")
            return None
        if not 0 <= col_no <self.get_max_col_count():
            print("输入的列号不在合法的列号范围内，需要大于等于0且小于最大列号的范围内")
            return None
        return self.sheet.cell(row=row_no+1,column=col_no+1).value

    #获取某一行（从0开始）的单元格的值，放到列表中
    def get_row_value(self,row_no):
        if (not isinstance(row_no,int)) :
            print("输入的行号%s不是整数" %(row_no))
            return  None
        if not 0<=row_no<self.get_max_row_count():
            print("输入的行号不在合法的行号范围内，需要大于等于0且小于最大行号的范围内")
            return None

        return self.get_sheet_all_cell_values()[row_no]


    # 获取某一列（从0开始）的单元格的值，放到列表中
    def get_col_value(self, col_no):
        if (not isinstance(col_no, int)):
            print("输入的列号%s不是整数" % (col_no))
            return None
        if not 0 <= col_no < self.get_max_col_count():
            print("输入的列号不在合法的列号范围内，需要大于等于0且小于最大列号的范围内")
            return None

        col_values = []
        for row in self.get_sheet_all_cell_values():
            for idx,cell in enumerate(row):
                if idx == col_no:
                    col_values.append(cell)
        return col_values


    #将多行数据写入到excel中，追加到excel 的最后面
    def write_lines_in_sheet(self,data,border_flag = True):
        if not isinstance(data,(list,tuple)):
            print("您写入的数据不是元组或列表类型，请重新设定")
            return
        for line in data:#data是需要包含多个子列表或子元组的类型
            if not isinstance(line, (list, tuple)):
                print("你写入的数据行,不是列表或者元组类型，请重新设定")
                return
            self.sheet.append(line)
        bd = Side(style='thin', color="000000")
        if border_flag:
            for row in self.sheet.rows:
                for cell in row:
                    cell.border = Border(left=bd,top=bd,right=bd,bottom=bd)
        self.save()

    #将一行内容追加到excel的最后面，设定是否有边框、字体的颜色（red、green，None--》黑色）、背景色
    def write_a_line_in_sheet(self,data,border_flag=True,font_color=None,fgcolor=None):
        if fgcolor is not None:
            fill = PatternFill(fill_type="solid",fgColor=fgcolor)
        else:
            fill = None
        if font_color is None:
            ft = None
        elif "red" in font_color:
            ft = Font(color=colors.RED)
        elif "green" in font_color:
            ft = Font(color=colors.GREEN)

        if not isinstance(data, (list, tuple)):
            print("你写入的数据行,不是列表或者元组类型，请重新设定")
            return
        first_line =self.sheet[1]
        if len(first_line)==1 and first_line[0].value is None:
            rowNo = self.get_max_row_count()
        else:
            rowNo = self.get_max_row_count() + 1

        for idx,value in enumerate(data):
            print(idx,value)
            if font_color is not None:
                #行号和列号都是从1开始
                #设定了一下填充的背景色
                #self.sheet.cell(row=rowNo,column=idx+1).fill = fill
                if ("成功" in str(value)) or ("失败" in str(value)) and ft is not None:
                    self.sheet.cell(row=rowNo, column=idx + 1).font = ft
            if fgcolor is not None:
                self.sheet.cell(row=rowNo, column=idx + 1).fill = fill
            self.sheet.cell(row=rowNo, column=idx + 1).value = value
        bd = Side(style='thin', color="000000")
        if border_flag:
            for row in self.sheet.rows:
                for cell in row:
                    cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.save()


    #写入某一列的值，列号从0开始
    def write_a_col_in_sheet(self,col_no,data,border_flag=True,font_color=None,fgcolor=None):
        if fgcolor is not None:
            fill = PatternFill(fill_type="solid", fgColor=fgcolor)
        else:
            fill = None
        if font_color is None:
            ft = None
        elif "red" in font_color:
            ft = Font(color=colors.RED)
        elif "green" in font_color:
            ft = Font(color=colors.GREEN)

        if not isinstance(data, (list, tuple)):
            print("你写入的数据列,不是列表或者元组类型，请重新设定")
            return

        bd = Side(style='thin', color="000000")
        for idx, value in enumerate(data):
            self.sheet.cell(row=idx + 1, column=col_no + 1).value = str(value)
            if font_color is not None:
                if ("成功" in str(value)) or ("失败" in str(value)) and ft is not None:
                    self.sheet.cell(row=idx + 1, column=col_no + 1).font = ft
            if fgcolor is not None:
                self.sheet.cell(row=idx + 1, column=col_no + 1).fill = fill
            if border_flag:
                self.sheet.cell(row=idx + 1, column=col_no + 1).border = Border(left=bd, top=bd, right=bd,
                                                                                    bottom=bd)

        self.save()


    #写一个单元格的值，行号从0开始，列号从0开始
    def write_cell_value(self, row_no, col_no, value, color=None, border=True):
        if (not isinstance(row_no, int)) or (not isinstance(col_no, int)):
            print("输入的行号%s,或者输入的列表%s不是整数" % (row_no, col_no))
            return None

        if color:
            if "red" in color:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.RED)
            elif "green" in color:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.GREEN)
            else:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.BLACK)
        if border:
            bd = Side(style='thin', color='000000')
            self.sheet.cell(row=row_no + 1, column=col_no + 1).border = Border(left=bd, top=bd, right=bd, bottom=bd)
        self.sheet.cell(row=row_no + 1, column=col_no + 1).value = value
        self.save()

    #在指定单元格写入年月日时间(行号从0开始，列号从0开始)
    def write_date_in_cell(self,row_no,col_num):
        #date = xxxxx.getCurrentDate()#获取当前时间的字符串
        from Util.DateStr import get_chinese_datetime
        date = get_chinese_datetime()
        #英文格式 date = get_english_datetime()
        #est时间 date = get_est_current_time()
        self.write_cell_value(row_no,col_num,date)

    #保存excel文件
    def save(self):
        self.wb.save(self.get_file_path())
if __name__ == "__main__":
    excel = ExcelUtil("d:\\phpStudy\\po_framework\\test.xlsx")
    """
    print(excel.get_sheet_names())
    
    print(excel.get_file_path())
    
    print(excel.set_sheet_by_index(3))
    print(excel.set_sheet_by_index(2))
    print(excel.set_sheet_by_index(1))
    
    print(excel.set_sheet_by_name("Sheetxx"))
    print(excel.set_sheet_by_name("Sheet1"))
    
    excel.create_sheet("测试数据")
    print(excel.get_sheet_names())
    
    excel.create_sheet("测试数据1")
    print(excel.get_sheet_names())
    
    excel.set_sheet_by_name("Sheet1")
    print(excel.get_max_col_count())
    print(excel.get_max_row_count())

    print(excel.set_sheet_by_name("Sheet1"))
    print(excel.get_sheet_all_cells())
    """

    print(excel.set_sheet_by_name("Sheet1"))
    print(excel.get_sheet_all_cell_values())

    """  
    print(excel.write_lines_in_sheet([['你好', "不错"], ['OK', "good"]]))
    print(excel.write_lines_in_sheet(['你好', "不错"]))
    

    print(excel.set_sheet_by_name("Sheet1"))
    excel.write_a_line_in_sheet(["你好", "wohao ", "ok", "成功"], font_color="red", fgcolor="CD9B9B")
    excel.write_a_line_in_sheet(["你好", "wohao ", "ok", "成功"], font_color="green", fgcolor="CD9B9B")
    
   
    excel.set_sheet_by_name("Sheet1")
    print(excel.get_cell_value(1, 2))
    

    excel.set_sheet_by_name("Sheet1")
    print(excel.get_row_value(2))
    print(excel.get_col_value(1))

    excel.set_sheet_by_name("Sheet1")
    excel.write_cell_value(14, 0, "write a cell", color="green", border=None)


    excel.set_sheet_by_name("测试数据")
    excel.write_a_col_in_sheet(3,[1,2,3,4,5],border_flag=True,font_color=None,fgcolor=None)
    excel.write_a_col_in_sheet(1, [1, 2, 3, 4, "失败"], True, "red")
    

    excel.set_sheet_by_name("测试数据")
    excel.write_date_in_cell(3,4)
    
    """